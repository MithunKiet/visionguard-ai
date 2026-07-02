"""
Report file generators — PDF (reportlab) and Excel (openpyxl). Pure
functions: rows in, file bytes out. Branding (enterprise name) is passed in
by the caller — never hardcoded (master context rule #1).
"""
import io
from datetime import datetime


def build_pdf(title: str, enterprise_name: str, period: str, headers: list[str], rows: list[list]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()

    elements = [
        Paragraph(f"{enterprise_name} — {title}", styles["Title"]),
        Paragraph(f"Period: {period}", styles["Normal"]),
        Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]),
        Spacer(1, 12),
    ]

    table_data = [headers] + (rows or [["No data for this period"] + [""] * (len(headers) - 1)])
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)

    doc.build(elements)
    return buf.getvalue()


def build_xlsx(title: str, enterprise_name: str, period: str, headers: list[str], rows: list[list]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    ws.append([f"{enterprise_name} — {title}"])
    ws.append([f"Period: {period}"])
    ws.append([])
    ws.append(headers)

    header_row = ws.max_row
    fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill

    for row in rows:
        ws.append(row)

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=header_row, column=col_idx).column_letter].width = max(14, len(header) + 4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
